import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import Alignment
import SqliteUtil
import json

mywb = openpyxl.Workbook()
mysheet = mywb.active   # 当前活动的sheet
boldFont = Font(size=12, bold=True)
textFont = Font(size=12)
alignCenter = Alignment(horizontal='center',vertical='center',wrap_text=True)
verticalCenter = Alignment(vertical='center',wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
currentRow = 1
orderInfo = {}
storeInfo= {}

def init():
    print('init')
    #mysheet.font = textFont
    #mysheet.column_dimensions['E'].font = textFont
    # 设置行列的宽高
    mysheet.row_dimensions[1].height = 60
    mysheet.column_dimensions['A'].width = 12
    mysheet.column_dimensions['B'].width = 28
    mysheet.column_dimensions['C'].width = 6
    mysheet.column_dimensions['D'].width = 6
    mysheet.column_dimensions['E'].width = 8
    mysheet.column_dimensions['F'].width = 12
    mysheet.column_dimensions['G'].width = 8
    mysheet.column_dimensions['H'].width = 8

def setBorder():
    rows = mysheet.iter_rows()
    rows = list(rows)[3:]
    for row in rows:
        for cell in row:
            cell.border = thin_border

def setCellBorder(cell):
    mysheet[cell].border = thin_border

def setText(cell, text, isBold=False, isCenter=True, setDefaultHeight=False):
    mysheet[cell] = text
    #mysheet[cell].border = thin_border
    if isCenter:
        mysheet[cell].alignment = alignCenter   
    else:
        mysheet[cell].alignment = verticalCenter  
    
    if isBold:
        mysheet[cell].font = boldFont
    else:
        mysheet[cell].font  = textFont
    
    # 设置该行默认行高
    if setDefaultHeight:
        setRowHeight(mysheet[cell].row, 20)


def setTextRange(cell_range, text, isBold=False, isCenter=True, setDefaultHeight=False):
    mysheet.merge_cells(cell_range)
    setText(cell_range.split(':')[0], text, isBold=isBold, isCenter=isCenter, setDefaultHeight=setDefaultHeight)

def setRowHeight(row, height):
    mysheet.row_dimensions[row].height = height


def buildExcel(id):
    global orderInfo
    global storeInfo
    global currentRow
    currentRow = 1
    orderInfo = SqliteUtil.getOrderById(id)
    storeInfo = SqliteUtil.getStore(1, isJson=False)
    init()
    buildTitle()
    buildClientInfo()
    buildStoreInfo()
    #mysheet.column_dimensions['E'].font = textFont
    
    buildRepairItems()
    buildParts()

    buildTotals()

    # 对前面已经构建好的内容设置边框，例如后面的footer就没有边框。
    #setBorder(mysheet, "A1:H10")
    setBorder()

    buildFooter()
    
    fileName = '%s.xlsx' % orderInfo['sn']
    mywb.save('./excel/%s' % fileName)
    print(fileName)
    return fileName

def buildTitle():
    global storeInfo
    global currentRow
    startRow = currentRow
    setTextRange('A1:H1', '%s维修清单' %(storeInfo['name']))
    mysheet['A1'].font = Font(size=18, bold=True)

    startRow+=1
    mysheet.merge_cells('A%d:E%d' % (startRow, startRow+1))

    setText('F%d' % startRow, '维修单号', setDefaultHeight=True)
    setText('F%d' % (startRow+1), '日期', setDefaultHeight=True)
    setTextRange('G%d:H%d' %(startRow, startRow), orderInfo['sn'], isCenter=False)
    setTextRange('G%d:H%d' %(startRow+1, startRow+1), orderInfo['create_time'].split(' ')[0], isCenter=False)
    setCellBorder('F%d' % startRow)
    setCellBorder('F%d' % (startRow+1))
    setCellBorder('G%d' % startRow)
    setCellBorder('G%d' % (startRow+1))
    setCellBorder('H%d' % startRow)
    setCellBorder('H%d' % (startRow+1))

    currentRow += 3

def buildClientInfo():
    global currentRow
    global orderInfo
    print('orderInfo: %s' % orderInfo['create_time'])
    startRow = currentRow
    
    """ 
    setTextRange('A%d:H%d' %(startRow, startRow), '', isBold=True)
    setRowHeight(startRow, 30)
    startRow+=1 
    """

    setText('A%d' % (startRow), '客户名称', setDefaultHeight=True)
    setTextRange('B%d:C%d' %(startRow, startRow), orderInfo['client_name'], isCenter=False)

    setTextRange('D%d:E%d' % (startRow, startRow), '联系人及电话')
    setTextRange('F%d:H%d' %(startRow, startRow), orderInfo['client_user'] + '  ' + orderInfo['client_phone'], isCenter=False)

    startRow+=1
    setText('A%d' % (startRow), '型号', setDefaultHeight=True)
    setTextRange('B%d:C%d' %(startRow, startRow), orderInfo['client_model'], isCenter=False)

    setTextRange('D%d:E%d' % (startRow, startRow), '车牌号码')
    setTextRange('F%d:H%d' %(startRow, startRow), orderInfo['client_sn'], isCenter=False)

    currentRow += 2


def buildStoreInfo():
    global storeInfo
    global currentRow
    startRow = currentRow

    #print(storeInfo)
    setTextRange('A%d:A%d' %(startRow, startRow+1), '地址/电话')
    setTextRange('B%d:H%d' %(startRow, startRow), storeInfo['address'], isCenter=False, setDefaultHeight=True)
    setTextRange('B%d:H%d' %(startRow+1, startRow+1), '联系人：%s    电话：%s' %(storeInfo['user'], storeInfo['phone']), isCenter=False, setDefaultHeight=True)

    currentRow += 2


def buildRepairItems():
    global currentRow
    global orderInfo
    startRow = currentRow
    setTextRange('A%d:H%d' %(startRow, startRow), '维修内容', isBold=True)
    setRowHeight(startRow, 30)

    startRow+=1
    setText('A%d' % (startRow), '序号', setDefaultHeight=True)
    setTextRange('B%d:E%d' % (startRow,startRow), '维修内容')
    setText('F%d' % (startRow), '维修费')
    setTextRange('G%d:H%d' % (startRow,startRow), '备注')

    repairItems = json.loads(orderInfo['repair_items'])
    index = 0
    total = 0
    for item in repairItems:
        #print(item)
        startRow+=1
        index+=1
        discount = 1 
        if('discount' in item):
            discount = item['discount']
        total+=(item['price'] * discount)
        setText('A%d' % startRow, '%d' % index, setDefaultHeight=True)
        setTextRange('B%d:E%d' % (startRow,startRow), item['name'])
        setText('F%d' % startRow, item['price'])
        setTextRange('G%d:H%d' % (startRow,startRow), '免费' if('discount' in item and item['discount']==0) else '')
    
    startRow+=1
    setTextRange('A%d:E%d' % (startRow,startRow), '合计')
    setText('F%d' % startRow, total, setDefaultHeight=True)
    setTextRange('G%d:H%d' % (startRow,startRow), '')

    currentRow += (3 + index)




def buildParts():
    global currentRow
    global orderInfo
    startRow = currentRow
    setTextRange('A%d:H%d' %(startRow, startRow), '材料费用', isBold=True, isCenter=True)
    setRowHeight(startRow, 30)

    startRow+=1
    setText('A%d' % (startRow), '序号', setDefaultHeight=True)
    setText('B%d' % (startRow), '材料名称')
    setText('C%d' % (startRow), '数量')
    setText('D%d' % (startRow), '单位')
    setText('E%d' % (startRow), '单价')
    setText('F%d' % (startRow), '金额')
    setTextRange('G%d:H%d' % (startRow,startRow), '备注')

    parts = json.loads(orderInfo['parts'])
    index = 0
    total = 0
    for part in parts:
        #print(part)
        startRow+=1
        index+=1
        discount = 1 
        if('discount' in part):
            discount = part['discount']
        total+=(part['price'] * part['count'] * discount)
        nameLength = len(part['name'])
        print(part['name'], nameLength)
        setText('A%d' % startRow, '%d' % index, setDefaultHeight=False if(nameLength > 18) else True)
        setText('B%d' % (startRow), part['name'])
        setText('C%d' % (startRow), part['count'])
        setText('D%d' % (startRow), part['unit'])
        setText('E%d' % (startRow), part['price'])
        setText('F%d' % (startRow), part['price'] * part['count'])
        setTextRange('G%d:H%d' % (startRow,startRow), '免费' if('discount' in part and part['discount']==0) else '')
    
    startRow+=1
    setTextRange('A%d:E%d' % (startRow,startRow), '合计')
    setText('F%d' % startRow, total, setDefaultHeight=True)
    setTextRange('G%d:H%d' % (startRow,startRow), '')

    currentRow += (3 + index)



def buildTotals():
    global currentRow
    global orderInfo
    startRow = currentRow
    setTextRange('A%d:H%d' %(startRow, startRow), '维修费用合计', isBold=True, isCenter=True)
    setRowHeight(startRow, 30)

    startRow+=1
    setText('A%d' % (startRow), '项目', setDefaultHeight=True)
    setTextRange('B%d:D%d' % (startRow,startRow), '小计')
    setTextRange('E%d:F%d' % (startRow,startRow), '优惠')
    setTextRange('G%d:H%d' % (startRow,startRow), '合计')

    totals = json.loads(orderInfo['totals'])
    index = 0
    total = 0
    for item in totals:
        print(item)
        startRow+=1
        index+=1
        total+=(item['price'] - item['discount'])
        setText('A%d' % startRow, item['name'], setDefaultHeight=True)
        setTextRange('B%d:D%d' % (startRow,startRow), item['price'])
        setTextRange('E%d:F%d' % (startRow,startRow), item['discount'])
    
    setTextRange('G%d:H%d' % (currentRow+2,startRow), total)
    currentRow += (2 + index)


def buildFooter():
    setRowHeight(currentRow, 40)
    setTextRange('A%d:E%d' % (currentRow,currentRow), '')
    setText('F%d' % currentRow, '客户：')
    setTextRange('G%d:H%d' % (currentRow,currentRow), '')

#buildExcel(1)
